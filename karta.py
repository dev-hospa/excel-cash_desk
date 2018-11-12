import openpyxl, sys


# store the header row in list called header
def create_header(sheet):
    last_col = sheet.max_column
    header = []
    for i in range(1, last_col + 1):
        header.append(sheet.cell(row=1, column=i).value)
    return header

# insert header line in the first row
def insert_header(new_sheet, header):
    index_col = 0
    for name in header:
        index_col += 1
        new_sheet.cell(row=1, column=index_col).value = name


# return the i-th position of column with account number
def find_account(sheet, name="Číslo účtu"):
    last_col = sheet.max_column
    for i in range(1, last_col + 1):
        value = sheet.cell(row=1, column=i).value
        if value == name:
            return i

# store the i-th position of row, where new account beginns
def bookmarks(sheet):
    account_col = find_account(sheet)   # number of column in which are account numbers
    last_row = sheet.max_row
    account_no = sheet.cell(row=2, column=account_col).value
    bookmarks = [2] # first row with account number
    for i in range(2, last_row + 1):
        value = sheet.cell(row=i, column=account_col).value
        if value != account_no:
            bookmarks.append(i)
            account_no = value
    bookmarks.append(last_row + 1)
    return bookmarks


def copy_row(orig_sheet, orig_row, new_sheet, new_row):
    last_col = orig_sheet.max_column
    new_row = new_row
    for i in range(1, last_col + 1):
        new_sheet.cell(row=new_row, column=i).value = orig_sheet.cell(row=orig_row, column=i).value



wb = openpyxl.load_workbook("261.xlsx")
journals = wb.active

header = create_header(journals)
new_wb = openpyxl.Workbook()


bookmark = bookmarks(journals)
for i in range(len(bookmark) - 1):
    new_sheet = new_wb.create_sheet(f"{bookmark[i]}")
    insert_header(new_sheet, header)
    new_row = 1
    for r in range(bookmark[i], bookmark[i+1]):
        new_row += 1
        copy_row(journals, r, new_sheet, new_row)


new_wb.save("new261.xlsx")
sys.exit

    
